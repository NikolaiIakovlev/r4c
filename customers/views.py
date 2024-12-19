import datetime
from django.http import HttpResponse
#Устанавливаем библиотеку openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment

from robots.models import Robot
from django.utils import timezone

from django.http import HttpResponse, HttpResponseBadRequest


def generate_excel_report(request):
    if request.method != 'GET':
        return HttpResponseBadRequest("Method not allowed.")
    
    try:
        start_date = timezone.now() - datetime.timedelta(days=7)
        robots = Robot.objects.filter(created__gte=start_date)

        workbook = Workbook()
        workbook.remove(workbook.active) # удаление первого листа

        models = robots.values_list('model', flat=True).distinct()
        for model in models:
            worksheet = workbook.create_sheet(title=model)
            worksheet.append(["Модель", "Версия", "Количество за неделю"])

            versions = robots.filter(model=model).values_list('version', flat=True).distinct()
            for version in versions:
                count = robots.filter(model=model, version=version).count()
                worksheet.append([model, version, count])

            #Форматирование
            for row in worksheet.iter_rows():
                for cell in row:
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                    cell.font = Font(bold=True)


        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=robot_report.xlsx'
        workbook.save(response)
        return response

    except Exception as e:
        return HttpResponseBadRequest(f"Error generating report: {str(e)}")

